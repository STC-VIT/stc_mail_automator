import openpyxl as xl
import os

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

def main():
    wb = xl.load_workbook("BREW PARTICIPANTS.xlsx")
    sheet = wb.active 


    # print(email.value, status.value)
    
    for i in range(2, sheet.max_row+1):
        name, email, status = sheet.cell(i, 2), sheet.cell(i, 3), sheet.cell(i, 4)
        if status.value == 'n': 
            res = send_mail('Welcome to BREW!', email.value)  
            if res == 202:
                sheet.cell(i, 5).value = 'Mail sent succesfully'
                print("\033[1;32mMail sent successfully!!")
            status.value = 'y'
        elif status.value == 'y':   
            print("\033[1;33mMail already sent")
            sheet.cell(i, 5).value = 'Mail already sent'

    wb.save("BREW PARTICIPANTS.xlsx")


def send_mail(sub, email_id):
    message = Mail(
        from_email='STC VIT<ramizkhan1199@outlook.com>',
        to_emails=email_id,
        subject=sub,
        html_content="""Hi There, Tech Enthusiast!<br><br>
We’re glad to be having you as a part of our Workshop-Brew. <strong>Brew is based on cloud computing</strong> and, to speed up the process and get kick started with our learning, We would urge you to sign up for the cloud service, which Brew will be based upon. <br>
<strong>The steps regarding the same are given below-</strong><br>
1.	Go to the Amazon Web Services home page.<br>
2.	Choose <strong>Create an AWS Account</strong>.<br>
<strong>Note</strong>: If you've signed in to AWS recently, the button might say <strong>Sign In to the Console</strong>.<br>
3.	Enter your account information, and then choose <strong>Continue</strong>.<br>
<strong>Important</strong>: Be sure that you enter your account information correctly, especially your email address. If you enter your email address incorrectly, you won't be able to access your account. If <strong>Create a new AWS account</strong> isn't visible, first choose <strong>Sign in to a different account</strong>, and then choose Create a new AWS account.<br>
4.	Choose <strong>Personal or Professional</strong>.<br>
<strong>Note</strong>: Personal accounts and professional accounts have the same features and functions.<br>
5.	Enter your company or personal information.<br>
6.	Read and accept the AWS Customer Agreement.<br>
<strong>Note</strong>: Be sure that you read and understand the terms of the AWS Customer Agreement.<br>
7.	Choose <strong>Create Account and Continue</strong>.<br>
You receive an email to confirm that your account is created. You can sign in to your new account using the email address and password you supplied. However, you can't use AWS services until you finish activating your account.<br>
It takes <strong>around 72 hours</strong>, for your account to be created. <br><br>
<strong>Note:</strong><br>
1.	We strongly recommend to not use the functions of AWS, without the guidance of an expert, as it can lead to the deduction of the corresponding service charge from your bank account. STC shall not be held responsible in case of such an event. <br>
2.	RuPay Cards are not accepted in AWS, We advise you to use alternate payment methods, in case of being a RuPay owner. <br><br>
If you have any doubts, we are always happy to help. <br><br>
Regards,<br> 
Team STC""")

    sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
    response = sg.send(message)
    # print(response.status_code, response.body, response.headers)
    # print(response.status_code)
    return response.status_code

# main()
send_mail('test', 'ramiz.cop@gmail.com')